import pyautogui
import openpyxl
import time
import pytesseract
import os
import tkinter as tk
import time
from PIL import Image
from tkinter import messagebox

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# No excel:
# B = Server name
# C = Item name
# D = Categoria
# E = Buy Offer
# F = Sell Offer

# X = 'arquivo'.png
x = ''

excel_path = r'D:\Programming\tibia_search\tibia_search\tibia_search_bot\tibia_search_db.xlsx'
image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)
sheet_name = 'DB'
#df = pd.read_excel(excel_path, sheet_name='DB', engine='openpyxl')

# Carregar a planilha usando openpyxl
wb = openpyxl.load_workbook(excel_path)
ws = wb['DB']

# posicao do mouse para o client do tibia
#mouse1x = 1163
#mouse1y = 313
mouse2x = 2522
mouse2y = 382
mouse3x = 859
mouse3y = 516
mouse4x = 865
mouse4y = 552

# Substitua as coordenadas pelos valores corretos para sua imagem
REGIAO_SELL_OFFER = (1340, 287, 1467, 302)  # Exemplo de coordenadas (left, upper, right, lower)
REGIAO_BUY_OFFER = (1340, 519, 1467, 532)  # Exemplo de coordenadas (left, upper, right, lower)

# pesquisa comum
items = ["Tibia Coins", "Gold Token", "Silver Token","Alicorn Ring", "Amulet of Theurgy", "Arboreal Crown", "Arboreal Ring", "Arboreal Tome", "Arcanomancer Folio", "Arcanomancer Regalia",
"Arcanomancer Sigil","Bloody Pincers", "Blueberry Cupcake", "Bone Fiddle", "Brimstone Fangs", "Broken Key Ring", "Broken Shamanic Staff", "Charged Alicorn Ring", "Charged Arboreal Ring",
"Charged Arcanomancer Sigil", "Charged Spiritthorn Ring", "Cobra Axe", "Cobra Boots", "Cobra Club", "Cobra Crossbow", "Cobra Hood", "Cobra Rod", "Cobra Sword", "Cobra Wand", 
"Cultish Mask", "Dawnfire Pantaloons", "Dawnfire Sherwani", "Decorative Ribbon","Dream Shroud", "Dwarven Armor", "Eldritch Bow", "Eldritch Breeches", "Eldritch Claymore", "Eldritch Cowl",
"Eldritch Cuirass", "Eldritch Folio", "Eldritch Greataxe", "Eldritch Hood", "Eldritch Quiver", "Eldritch Rod", "Eldritch Shield", "Eldritch Tome", "Eldritch Wand", "Eldritch Warmace", "Elven Hoof", 
"Elven Mail", "Elvish Talisman", "Enchanted Pendulet", "Enchanted Sleep Shawl", "Enchanted Theurgic Amulet", "Enchanted Turtle Amulet", "Exalted Core", "Exotic Amulet", "Exotic Legs", "Fabulous Legs",
"Falcon Battleaxe", "Falcon Bow", "Falcon Circlet", "Falcon Coif", "Falcon Escutcheon", "Falcon Greaves", "Falcon Longsword","Falcon Mace", "Falcon Plate", "Falcon Rod", "Falcon Shield", 
"Falcon Wand", "Feverbloom Boots", "Final Judgement", "Four-leaf Clover", "Foxtail", "Frostflower Boots", "Galea Mortis", "Ghost Chestplate", "Giant Shrimp", "Gloom Wolf Fur", "Gnome Armor", 
"Gnome Helmet", "Gnome Legs", "Golden Can of Oil", "Gooey Mass", "Goosebump Leather", "Green Dragon Leather", "Library Ticket", "Lion Amulet", "Lion Axe", "Lion Hammer", "Lion Longbow", 
"Lion Plate", "Lion Rod", "Lion Shield", "Lion Spangenhelm", "Lion Spellbook", "Lion Wand", "Little bowl of Myrrh", "Make-do Boots", "Midnight Sarong", "Midnight Tunic", 
"Minor Crystalline Token", "Mooh'tah Shell", "Moohtant Horn", "Mutant Bone Boots", "Mutant Bone Kilt", "Mutated Skin Armor", "Mutated Skin Legs", "Naga Axe", "Naga Club", "Naga Crossbow", 
"Naga Quiver", "Naga Rod", "Naga Wand", "Nightmare Horn", "Ornate Chestplate", "Pair of Dreamwalkers", "Pair of Soulstalkers", "Pair of Soulwalkers", "Perfect Behemoth Fang", 
"Piece of Dead Brain", "Piece of Swampling Wood", "Prismatic Ring", "Protective Charm", "Rope Belt", "Rotworm Stew", "Sabretooth", "Sanguine Boots", "Sanguine Galoshes", "Sanguine Greaves", 
"Sanguine Legs", "Scrubbing Brush","Shadow Cowl", "Shoulder Plate", "Silencer Claws",  "Sliver", "Slug Drug", "Snake Skin", "Soap", "Some Grimeleech Wings", "Soulbastion", 
"Soulbleeder", "Soulcrusher", "Soulcutter", "Souleater", "Soulhexer", "Soulmaimer", "Soulmantle", "Soulpiercer", "Soulshanks", "Soulshell", "Soulshredder", "Soulshroud", "Soulstrider", 
"Soultainter", "Spirit Guide", "Spiritthorn Armor", "Spiritthorn Helmet", "Stitched Mutant Hide Legs", "Strand of Medusa Hair", "Strawberry Cupcake", "Suspicious Device", 
"Sweet Mangonaise Elixir", "The Cobra Amulet", "Thick Fur", "Toga Mortis", "Turtle Shell", "Umbral Master Spellbook", "Vampire Teeth", "Winter Wolf Fur", "Wyvern Talisman"]

# pesquisa o segundo item (item grand sanguine e enchanted aparece primeiro)
items_grand_sanguine = ["Pendulet", "Sleep Shawl", "Turtle Amulet", "Spiritthorn Ring", "Sanguine Battleaxe", "Sanguine Blade",  "Sanguine Bludgeon", "Sanguine Bow", "Sanguine Coil", "Sanguine Crossbow",
"Sanguine Cudgel", "Sanguine Hatchet", "Sanguine Razor", "Sanguine Rod"]

aux = 0
aux_sanguine = 0

# preencher com servidor
print(f"Informe o Servidor: ")
identificador_servidor = input() 

#functions ---------------------------------------------------------------------------------------------------------------------------
def exibir_caixa_mensagem():
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Programa Finalizado", "O programa foi concluído. Clique OK para confirmar.")

def clicar(posicao):
    pyautogui.click(posicao)

def digitar(texto):
    pyautogui.write(texto)

def BuscarItem():
    global items, identificador_item
    identificador_item = items[aux]

    #pyautogui.rightClick((mouse1x, mouse1y))
    
    #time.sleep(1)
    
    pyautogui.rightClick((mouse2x, mouse2y))
    
    time.sleep(1)
    
    digitar(identificador_item)
    
    time.sleep(1)
    
    pyautogui.click((mouse3x, mouse3y))
    
    time.sleep(1)
    
    pyautogui.press('=')  
    
    time.sleep(1)
    
    pyautogui.press('esc')
    
    #time.sleep(1)
    
    #pyautogui.rightClick((mouse1x, mouse1y))

def BuscarItemSanguine():
    global items_grand_sanguine, identificador_item_grand_sanguine
    identificador_item_grand_sanguine = items_grand_sanguine[aux]

    #pyautogui.rightClick((mouse1x, mouse1y))
    
    #time.sleep(1)
    
    pyautogui.rightClick((mouse2x, mouse2y))
    
    time.sleep(1)
    
    digitar(identificador_item_grand_sanguine)
    
    time.sleep(1)
    pyautogui.click((mouse4x, mouse4y))
    
    time.sleep(1)
    
    pyautogui.press('=')  
    
    time.sleep(1)
    
    pyautogui.press('esc')
    
    #time.sleep(1)
    
    #pyautogui.rightClick((mouse1x, mouse1y))

def achar_servidor_e_item():
    global cell_sell_offer, cell_buy_offer, identificador_item, identificador_servidor, excel_path, ws, wb
    
    # Iterar sobre as linhas da planilha para encontrar a posição
    servidor_col_idx = 1  # Coluna B (0-indexed seria 1)
    item_col_idx = 2      # Coluna C (0-indexed seria 2)
    
    servidor_row = None
    item_row = None

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        if row[servidor_col_idx].value == identificador_servidor:
            servidor_row = row[servidor_col_idx].row
            if row[item_col_idx].value == identificador_item:
                item_row = row[item_col_idx].row
                break
    
    # Verificar se encontramos algum resultado
    if item_row:
        print(f"Servidor '{identificador_servidor}' com item '{identificador_item}' encontrado na linha {item_row}.")
        # Definir a coluna
        coluna_sell_offer_letra = 'E'
        cell_sell_offer = f"{coluna_sell_offer_letra}{item_row}"
        coluna_buy_offer_letra = 'F'
        cell_buy_offer = f"{coluna_buy_offer_letra}{item_row}"
        print(cell_sell_offer)
        print(cell_buy_offer)
    else:
        print(f"Servidor '{identificador_servidor}' com item '{identificador_item}' não encontrado.")

def achar_servidor_e_item_sanguine():
    global cell_buy_offer, cell_sell_offer, identificador_item_grand_sanguine, identificador_servidor, excel_path, ws, wb
    
    # Iterar sobre as linhas da planilha para encontrar a posição
    servidor_col_idx = 1  # Coluna B (0-indexed seria 1)
    item_col_idx = 2      # Coluna C (0-indexed seria 2)
    
    servidor_row = None
    item_row = None

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        if row[servidor_col_idx].value == identificador_servidor:
            servidor_row = row[servidor_col_idx].row
            if row[item_col_idx].value == identificador_item_grand_sanguine:
                item_row = row[item_col_idx].row
                break
    
    # Verificar se encontramos algum resultado
    if item_row:
        print(f"Servidor '{identificador_servidor}' com item '{identificador_item_grand_sanguine}' encontrado na linha {item_row}.")
        # Definir a coluna
        coluna_sell_offer_letra = 'E'
        cell_sell_offer = f"{coluna_sell_offer_letra}{item_row}"
        coluna_buy_offer_letra = 'F'
        cell_buy_offer = f"{coluna_buy_offer_letra}{item_row}"
        print(cell_sell_offer)
        print(cell_buy_offer)
    else:
        print(f"Servidor '{identificador_servidor}' com item '{identificador_item_grand_sanguine}' não encontrado.")

def preencher_valor():
    # localizar a linha exatada do itemXservidor e preencher os valores no preço de compra e preço de venda
    global cell_buy_offer, cell_sell_offer, sell_offer, buy_offer, excel_path, wb, ws

    print(f"sell_offer inicial: {sell_offer}")
    print(f"buy_offer inicial: {buy_offer}")

    # Verificar se as variáveis globais estão definidas corretamente
    if not all([cell_buy_offer, cell_sell_offer, sell_offer, buy_offer, excel_path]):
        print("Uma ou mais variáveis globais não estão definidas corretamente.")
        return
    
    # Verificar e definir valores padrão caso não sejam encontrados
    if sell_offer == '' or sell_offer == '0' or sell_offer == None or sell_offer == 'None':
        print("sell_offer não encontrado, definindo como 0.")
        sell_offer = '0'
    if buy_offer == '' or buy_offer == '0' or buy_offer == None or buy_offer == 'None':
        print("buy_offer não encontrado, definindo como 0.")
        buy_offer = '0'
   
    print(f"sell_offer inicial: {sell_offer}")
    print(f"buy_offer inicial: {buy_offer}")

    # Preencher as celulas
    ws[cell_sell_offer] = int(sell_offer)
    ws[cell_buy_offer] = int(buy_offer)

    print(f"Valores preenchidos: Preço de Compra na célula {cell_sell_offer} e Preço de Venda na célula {cell_buy_offer}.")
    print(f"")

def remover_arquivos(file_path):
    global x
    os.remove(file_path)
    print(f"Arquivo {file_path} removido com sucesso.")
    print(f"")
    x = ''

def pegar_nome_arquivo_png():
    global x, image_path
    # Listar todos os arquivos na pasta especificada
    arquivos = os.listdir(image_path)
    
    # Procurar o primeiro arquivo com a extensão .png
    for arquivo in arquivos:
        if arquivo.endswith('.png'):
            x = arquivo  # Atribuir o nome do arquivo PNG à variável global x
            print(x)
            return 
               
def extrair_valor_img(regiao):
    global x
    # atualiza o caminho da imagem com o valor de X
    image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)

    # Carregar a imagem
    img = Image.open(image_path)
    
    # Cortar a região da imagem
    area = img.crop(regiao)
    
    # Extrair texto da região
    texto = pytesseract.image_to_string(area, lang='eng')
    
    # Procurar o primeiro valor numérico no texto extraído
    for word in texto.split():
        try:
            # Remover pontos e vírgulas
            numero = word.replace(',', '').replace('.', '').replace('$', '')
            return int(numero)
        except ValueError:
            continue
    
    return None 

def salvar_db():
    global wb

    # Salvar as mudanças
    wb.save(excel_path)

def mainItem():

    global item, sell_offer, buy_offer, x, aux
    num_items = len(items)

    while aux < num_items:
        item = items[aux]
        BuscarItem()
        pegar_nome_arquivo_png()
        achar_servidor_e_item()
        sell_offer = str(extrair_valor_img(REGIAO_SELL_OFFER))
        buy_offer = str(extrair_valor_img(REGIAO_BUY_OFFER))
        preencher_valor()
        image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)
        remover_arquivos(image_path)
        image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)
        x = ''
        aux += 1

def mainItemSanguine():
    global item_grand_sanguine, sell_offer, buy_offer, x, aux
    num_items_grand_sanguine = len(items_grand_sanguine)
    aux = 0 # usa o mesmo aux do mainItem e reseta ao iniciar a função

    while aux < num_items_grand_sanguine:
        item_grand_sanguine = items_grand_sanguine[aux]
        BuscarItemSanguine()
        pegar_nome_arquivo_png()
        achar_servidor_e_item_sanguine()
        sell_offer = str(extrair_valor_img(REGIAO_SELL_OFFER))
        buy_offer = str(extrair_valor_img(REGIAO_BUY_OFFER))
        preencher_valor()
        image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)
        remover_arquivos(image_path)
        image_path = os.path.join(r'C:\Users\joaov\AppData\Local\Tibia\packages\Tibia\screenshots',x)
        x = ''
        aux += 1

#test -----------------------------------------------------------------------------------------------------------------------------------

# testar a posicao do mouse
#posicao_mouse = pyautogui.position()
#print(f"A posição atual do mouse é: {posicao_mouse}")

#execution -------------------------------------------------------------------------------------------------------------------------------
start_time = time.time()
time.sleep(5)
mainItem()
mainItemSanguine()
salvar_db()
end_time = time.time()
execution_time = end_time - start_time
exibir_caixa_mensagem()
print(f"Tempo de execução: {execution_time:.2f} segundos")

# configuracoes da janela do tibia --------------------------------------------------------------------------------------------------------
# - apenas local chat e log abertos
# - 1 barra lateral na esquerda e uma na direita
# - Barra de vida no topo com barra de XP
# - Control buttons minimizado


"""
TO DO
* Implementar quantidade de itens/ofertas no market 
    - da mesma forma que o programa pega o preço dos itens usando a imagem ele pode pegar a quantidade
    de itens a venda em coluna e somar toda a quantidade, o lado negativo é que a quantidade de itens
    não irá bater com a quantidade de itens com o preço presente no db
* Implementar função para calculo de profit entre compra e venda de itens entre servidores 
    - listar quais e quantos itens estão disponíves para compra e venda entre os 2 servidores escolhidos,
    calcular o retorno monetário "ROI"
    - feito no powerBI (?)
* Implementar função de pesquisa de items com Tier (T1 - T10)
    - muito trabalho, 10 linhas no banco de dados para cada item classificação 4, tirando os outros itens
    com classificação inferior
"""
